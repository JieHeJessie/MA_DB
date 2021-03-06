import re

from django.shortcuts import render
from django.urls import reverse
from django.shortcuts import redirect
from migration_db.models import Project, Node,quote,user,invoice
# Create your views here.
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.db.models import Sum, Q
import openpyxl
from MA_DB import settings
import os

import MySQLdb
from datetime import date, datetime
import decimal
STATES = ['Vic','Qld','NSW','SA','WA','NT']
def hello(request):
    return  HttpResponse("Hello world")


def login(request):

    return render(request, 'login.html')


def login_result(request):
    context_dict = {}
    if request.method=='POST':
        username = request.POST['username']
        password = request.POST['pass_word']

        if user.objects.filter(username=username).filter(password=password).exists():
            request.session['username']=username
            return redirect('index')
        else:
            context_dict['msg']='msg'
            ret_msg=u'Your username or password is not correct.Please log in again'
            return render(request, 'login.html',{'msg': ret_msg})

def register(request):
    return render(request,'register.html')

def register_result(request):
    context_dict = {}
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        if user.objects.filter(username=username).exists():
            context_dict['status'] = 'fail'
        else:
            context_dict['status']='success'
            user.objects.create(username=username, password=password)
        return render(request, 'register_result.html', context_dict)




def import_page(request):
    return render(request, 'import.html')

def import_project(request):
    return render(request, 'import_project.html')

def import_invoice(request):
    return render(request, 'import_invoice.html')


def import_quote(request):
    return render(request, 'import_quote.html')
def read_project(request):
    context_dict = {}
    if request.method == 'POST':
        file_name = request.FILES['excel']
        print request.FILES['excel']
        sheet=request.POST['sheet']
        start=int(request.POST['from_row'])
        end=int(request.POST['to_row'])

        book = openpyxl.load_workbook(file_name, data_only=True)
        sheet = book.get_sheet_by_name(sheet)

        database = MySQLdb.connect(host="localhost", user=settings.DB_USERNAME, passwd=settings.DB_PASSWORD, db="ma_newdb")
        cursor = database.cursor()

        node_header = sheet['A' + str(1)].value
        pro_date_header = sheet['B' + str(1)].value
        project_name_header = sheet['C' + str(1)].value
        service_header = sheet['D' + str(1)].value
        instrument_header = sheet['E' + str(1)].value
        person_header = sheet['F' + str(1)].value
        organization_header = sheet['G' + str(1)].value
        num_sample_header = sheet['H' + str(1)].value
        category_header = sheet['I' + str(1)].value
        int_ext_header = sheet['J' + str(1)].value
        state_header = sheet['K' + str(1)].value
        country_header = sheet['L' + str(1)].value
        user_define1_header = sheet['M' + str(1)].value
        user_define2_header = sheet['N' + str(1)].value
        subtotal_header = sheet['O' + str(1)].value
        cus_count_header = sheet['P' + str(1)].value
        if node_header!='Node':
            context_dict['state'] = 'fail'
            context_dict['table']='project'
            context_dict['error_message']="the conlumn A name should be 'Node'"
            return render(request, 'import_success.html',context_dict)
        if pro_date_header!='Date':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn B name should be 'Date'"
            return render(request, 'import_success.html',context_dict)
        if project_name_header!='Project':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn C name should be 'Project'"
            return render(request, 'import_success.html',context_dict)
        if service_header!='Service':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn D name should be 'Service',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if instrument_header!='Instrument':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn E name should be 'Instrument',,please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if person_header!='Person':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn F name should be 'Person',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if organization_header!='Person/Organisation':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn G name should be 'Person/Organisation',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if num_sample_header!='No. samples*':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn H name should be 'No. samples*',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if category_header!='Category':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn I name should be 'Category',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if int_ext_header!='Int/Ext':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn J name should be 'Int/Ext',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if state_header!='State':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn K name should be 'State',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if country_header!='Country':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn L name should be 'Country',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if user_define1_header!='User Definined 1':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn M name should be 'User Definined 1',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if user_define2_header!='User Definined 2':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn N name should be 'User Definined 2',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if subtotal_header!='Sub-total':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn O name should be 'Sub-total',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        if cus_count_header!='Customer Number':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'project'
            context_dict['error_message']="the conlumn P name should be '',please check spreadsheet format notice."
            return render(request, 'import_success.html',context_dict)
        value_list = []
        index = 0
        for row in range(start, end+1):
            node = sheet['A' + str(row)].value
            pro_date = sheet['B' + str(row)].value
            project_name = sheet['C' + str(row)].value
            service = sheet['D' + str(row)].value
            instrument = sheet['E' + str(row)].value
            person = sheet['F' + str(row)].value
            organization = sheet['G' + str(row)].value
            num_sample = int(sheet['H' + str(row)].value)
            category = sheet['I' + str(row)].value
            int_ext = sheet['J' + str(row)].value
            state = sheet['K' + str(row)].value
            country = sheet['L' + str(row)].value
            user_define1 = sheet['M' + str(row)].value
            user_define2 = sheet['N' + str(row)].value
            subtotal = sheet['O' + str(row)].value
            cus_count = sheet['P' + str(row)].value

            columns='''ABCDEFGHIJKLMNOP'''
            all_none=True
            for c in columns:
                if sheet[c+str(row)].value!=None:
                    all_none=False
            if all_none:
                index += 1
                continue
            if not re.match('^[A-Za-z]+$', node):
                context_dict['state'] = 'fail'
                context_dict['error_message'] = "invalid node at %d %s" % (index + start, 'A')
                context_dict['node']=node
                context_dict['table']='project'
                return render(request, 'import_success.html', context_dict)
            if type(pro_date)!=type(datetime.strptime('2010-01-01','%Y-%m-%d')):
                context_dict['state'] = 'fail'
                print type(pro_date)
                print pro_date
                context_dict['error_message'] = "invalid date at %d %s" % (index + start, 'B')
                context_dict['table']='project'
                return render(request, 'import_success.html', context_dict)
            if not re.match('^\d{1,2}/\d{1,2}/\d{4}$', datetime.strftime(pro_date,'%d/%m/%Y')):
                context_dict['state'] = 'fail'
                context_dict['table']='project'
                context_dict['error_message'] = "invalid Date at %d %s" % (index + start, 'B')
                return render(request, 'import_success.html', context_dict)
            index += 1
            # if pro_date=='Half year' or pro_date is 'Full year':
            # pro_date = date.today().isoformat()
        for row in range(start, end + 1):
            node = sheet['A' + str(row)].value
            pro_date = sheet['B' + str(row)].value
            project_name = sheet['C' + str(row)].value
            service = sheet['D' + str(row)].value
            instrument = sheet['E' + str(row)].value
            person = sheet['F' + str(row)].value
            organization = sheet['G' + str(row)].value
            num_sample = int(sheet['H' + str(row)].internal_value)
            category = sheet['I' + str(row)].value
            int_ext = sheet['J' + str(row)].value
            state = sheet['K' + str(row)].value
            country = sheet['L' + str(row)].value
            user_define1 = sheet['M' + str(row)].value
            user_define2 = sheet['N' + str(row)].value
            subtotal = sheet['O' + str(row)].value
            cus_count = sheet['P' + str(row)].value
        if isinstance(pro_date, date):
            pro_date = pro_date.isoformat()
        else:
            pro_date = date.today().isoformat()
        if project_name is None:
            project_name = 'null'
        if person is None:
            person = 'Null'
        if organization is None:
            organization = "Null"
        if cus_count is None:
            cus_count = "0"
        if user_define2 is not None:
            user_define2 = user_define2.upper()
        else:
            user_define2 = "NULL"
        if not isinstance(num_sample, int):
            num_sample = 0

        if subtotal is None:
            subtotal = 0.00
        else:
            subtotal = decimal.Decimal("%.2f" % subtotal)

        field_check = user_define2.split(';')
        if len(field_check) == 1:
            query = """INSERT INTO migration_db_project (node, pro_date,project_name,service,instrument,person,organization,num_sample,category,int_ext,state,country,user_define1,user_define2,subtotal,cus_count) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            values = (
                node, pro_date, project_name, service, instrument, person, organization, num_sample, category,
                int_ext, state,
                country, user_define1, user_define2, subtotal, cus_count)
            cursor.execute(query, values)
            database.commit()
        else:
            for each in field_check:
                query = """INSERT INTO migration_db_project (node,pro_date,project_name,service,instrument,person,organization,num_sample,category,int_ext,state,country,user_define1,user_define2,subtotal,cus_count) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
                values = (
                    node, pro_date, project_name, service, instrument, person, organization, num_sample, category,
                    int_ext,
                    state, country, user_define1, each, subtotal, cus_count)
                cursor.execute(query, values)
                database.commit()

        cursor.close()

        database.close()

        print ("Importion Done.")
        context_dict['state'] = 'success'
        return render(request, 'import_success.html',context_dict)

def read_invoice(request):
    context_dict = {}
    if request.method == 'POST':
        file_name = request.FILES['excel']
        print request.FILES['excel']
        sheet=request.POST['sheet']
        start=int(request.POST['from_row'])
        end=int(request.POST['to_row'])

        book = openpyxl.load_workbook(file_name, data_only=True)
        sheet = book.get_sheet_by_name(sheet)

        database = MySQLdb.connect(host="localhost", user=settings.DB_USERNAME, passwd=settings.DB_PASSWORD, db="ma_newdb")
        cursor = database.cursor()

        date_header = sheet['A' + str(1)].value
        inv_header = sheet['B' + str(1)].value

        quote_header = sheet['D' + str(1)].value
        paid_header = sheet['E' + str(1)].value
        ma_staff_header = sheet['F' + str(1)].value
        project_inv_header = sheet['G' + str(1)].value
        service_type_header = sheet['H' + str(1)].value
        instrument_header = sheet['I' + str(1)].value
        person_header = sheet['J' + str(1)].value
        address_header = sheet['K' + str(1)].value
        no_sample_header = sheet['L' + str(1)].value
        category_header = sheet['M' + str(1)].value
        int_ext_header = sheet['N' + str(1)].value
        user_define1_header = sheet['O' + str(1)].value
        user_define2_header = sheet['P' + str(1)].value
        sub_total_header = sheet['Q' + str(1)].value
        index = 0
        if date_header != 'Date':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Date'"
        if inv_header != 'Inv#':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Inv#'"
        if quote_header != 'Quote#':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Quote#'"
        if paid_header != 'Paid':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Paid'"
        if ma_staff_header != 'MA Staff':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'MA Staff'"
        if project_inv_header != 'Project':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Project'"
        if service_type_header != 'Type of Service':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Type of Service'"
        if instrument_header != 'Instrument':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Instrument'"
        if person_header != 'Person':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Person'"
        if address_header != 'Address':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Address'"
        if no_sample_header != 'No.Sample':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'No.Sample'"
        if category_header != 'Category':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Category'"
        if int_ext_header != 'Int/Ext':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Int/Ext'"
        if user_define1_header != 'User Defined 1':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'User Defined 1'"
        if user_define2_header != 'User Defined 2':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'User Defined 2'"
        if sub_total_header != 'Sub-Total':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'invoice'
            context_dict['error_message'] = "the conlumn A name should be 'Sub-Total'"

        for row in range(start, end + 1):
            invoice_date = sheet['A' + str(row)].value
            invoice_num = sheet['B' + str(row)].value
            bad_debt = sheet['C' + str(row)].value
            quote = sheet['D' + str(row)].value
            paid = sheet['E' + str(row)].value
            ma_staff = sheet['F' + str(row)].value
            project_invoice = sheet['G' + str(row)].value
            service_type = sheet['H' + str(row)].value
            instrument = sheet['I' + str(row)].value
            person_invoice = sheet['J' + str(row)].value
            address = sheet['K' + str(row)].value
            no_sample_in = sheet['L' + str(row)].value
            category_in = sheet['M' + str(row)].value
            int_ext_in = sheet['N' + str(row)].value
            user_define1_in = sheet['O' + str(row)].value
            user_define2_in = sheet['P' + str(row)].value
            sub_total_in = sheet['Q' + str(row)].value
            reconciliation = sheet['R' + str(row)].value
            running_total = sheet['S' + str(row)].value
            columns = '''ABCDEFGHIJKLMNOPQRS'''
            all_none = True
            for c in columns:
                if sheet[c + str(row)].value != None:
                    all_none = False
            if all_none:
                index += 1
                continue

            if type(invoice_date) != type(datetime.strptime('2010-01-01', '%Y-%m-%d')):
                context_dict['state'] = 'fail'
                print type(invoice_date)
                print invoice_date
                context_dict['error_message'] = "invalid date at %d %s" % (index + start, 'B')
                context_dict['table'] = 'project'
                return render(request, 'import_success.html', context_dict)




        for row in range(start, end+1):
            invoice_date = sheet['A' + str(row)].value
            invoice_num = sheet['B' + str(row)].value
            bad_debt = sheet['C' + str(row)].value
            quote = sheet['D' + str(row)].value
            paid = sheet['E' + str(row)].value
            ma_staff = sheet['F' + str(row)].value
            project_invoice = sheet['G' + str(row)].value
            service_type = sheet['H' + str(row)].value
            instrument = sheet['I' + str(row)].value
            person_invoice = sheet['J' + str(row)].value
            address = sheet['K' + str(row)].value
            no_sample_in = sheet['L' + str(row)].value
            category_in = sheet['M' + str(row)].value
            int_ext_in = sheet['N' + str(row)].value
            user_define1_in = sheet['O' + str(row)].value
            user_define2_in = sheet['P' + str(row)].value
            sub_total_in = sheet['q' + str(row)].value
            reconciliation = sheet['r' + str(row)].value
            running_total = sheet['s' + str(row)].value
            if isinstance(invoice_date, date):
                invoice_date = invoice_date.isoformat()
            else:
                invoice_date = date.today().isoformat()
            if bad_debt is None:
                bad_debt = "Null"
            if paid is None:
                paid = "Null"
            if reconciliation is None:
                reconciliation = "Null"
            if running_total is None:
                running_total = "Null"

            query = """INSERT INTO migration_db_invoice(invoice_date,invoice_num,bad_debt,quote,paid,ma_staff
                                                            ,project_invoice,service_type,instrument,person_invoice,address,no_sample_in,category_in,
                                                            int_ext_in, user_define1_in, user_define2_in, sub_total_in, reconciliation, running_total) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            print query
            values = (invoice_date, invoice_num, bad_debt, quote, paid, ma_staff, project_invoice, service_type,
                      instrument, person_invoice, address, no_sample_in, category_in, int_ext_in, user_define1_in,
                      user_define2_in, sub_total_in, reconciliation, running_total)
            cursor.execute(query, values)

        database.commit()
        cursor.close()

        database.close()
        print ("Importion Done.")

        return render(request, 'import_success.html')

def read_quote(request):
    context_dict={}
    if request.method == 'POST':
        file_name = request.FILES['excel']
        print request.FILES['excel']
        sheet=request.POST['sheet']
        start=int(request.POST['from_row'])
        end=int(request.POST['to_row'])

        book = openpyxl.load_workbook(file_name, data_only=True)
        sheet = book.get_sheet_by_name(sheet)

        database = MySQLdb.connect(host="localhost", user=settings.DB_USERNAME, passwd=settings.DB_PASSWORD, db="ma_newdb")
        cursor = database.cursor()

        value_list=[]

        index=0
        quote_header = sheet['A' + str(1)].value
        year_header = sheet['B' + str(1)].value

        version_header = sheet['C' + str(1)].value
        concatenate_header = sheet['D' + str(1)].value
        client_header = sheet['E' + str(1)].value
        company_header = sheet['F' + str(1)].value
        mastaff_header = sheet['G' + str(1)].value
        date_header = sheet['H' + str(1)].value
        value_header = sheet['I' + str(1)].value
        grant_header = sheet['J' + str(1)].value
        accept_header = sheet['K' + str(1)].value
        invoice_header = sheet['L' + str(1)].value
        comment_header = sheet['M' + str(1)].value

        index = 0
        if quote_header != 'Quote #':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Quote #'"
        if year_header != 'Year':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Year'"
        if version_header != 'Version':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Version'"
        if concatenate_header != 'CONCATENATE':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'CONCATENATE'"
        if client_header != 'Client':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Client'"
        if company_header != 'Company':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Company'"
        if mastaff_header != 'MA Staff Member':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'MA Staff Member'"
        if date_header != 'Date':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Date'"
        if value_header != 'Value (incl. GST)':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Value (incl. GST)'"
        if grant_header != 'For grant':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'For grant'"
        if accept_header != 'Accepted':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Accpeted'"
        if invoice_header != 'Invoiced':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Invoiced'"
        if comment_header != 'Comment':
            context_dict['state'] = 'fail'
            context_dict['table'] = 'quote'
            context_dict['error_message'] = "the conlumn A name should be 'Comment'"




        for row in range(start, end+1):
            quote_num = sheet['A' + str(row)].value
            quote_year = sheet['B' + str(row)].value
            version = sheet['C' + str(row)].value
            concatenate = sheet['D' + str(row)].value
            client = sheet['E' + str(row)].value
            company = sheet['F' + str(row)].value
            quote_staff = sheet['G' + str(row)].internal_value
            quote_date = sheet['H' + str(row)].value
            quote_value = sheet['I' + str(row)].value
            grant_not = sheet['J' + str(row)].value
            accept = sheet['K' + str(row)].value
            invoiced_not = sheet['L' + str(row)].value
            comment = sheet['M' + str(row)].value
            if not re.match('',quote_num):
                errot_message="invalid quote number at %d %s"%(index+start,'A')
            if isinstance(quote_date, date):
                quote_date = quote_date.isoformat()
            else:
                quote_date = date.today().isoformat()

            if quote_value is None:
                quote_value = "Null"
            if version is None:
                version = "Null"

            if grant_not is None:
                grant_not = "Null"
            if accept is None:
                accept = "Null"
            if invoiced_not is None:
                invoiced_not = "Null"
            if comment is None:
                comment = "Null"
            index+=1
            query = """INSERT INTO migration_db_quote(quote_num,quote_year,version,concatenate,client,company,quote_staff,quote_date,quote_value,grant_not,accept, invoiced_not,comment) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            values = (
            quote_num, quote_year, version, concatenate, client, company, quote_staff, quote_date, quote_value,
            grant_not, accept, invoiced_not, comment)
            cursor.execute(query, values)
            database.commit()

        cursor.close()

        if errot_message!="":
            return
        database.close()

        print ("Importion Done.")

        return render(request, 'import_success.html')



def import_invoice(request):
    return render(request, 'import_invoice.html')

def import_quote(request):
    return render(request, 'import_quote.html')
# def import_project(request):
#     return render()

def chart_page(request):
    return render(request,'allNodesByNode.html')

def state_page(request):
    return render(request,'allNodesByState.html')

def compute_ab(request):
    return render(request,'compute_ab.html')

def index(request):
    if request.session.get("username","") =="":
        return redirect("login")
    return render(request,'index.html')

def filter_page(request,node_name_slug):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    context_dict['node_name'] = node_name_slug
    return render(request,'nodeFilter.html',context_dict)


def chartbynode(request):
    if request.session.get("username","") =="":
        return redirect("login")
    if request.method == 'POST':
        from_date = request.POST['from_user_date']
        to_date = request.POST['to_user_date']
        num_cus_va = None
        subtotal_va = None
        if 'num_cus' in request.POST:
            num_cus_va = request.POST['num_cus']
        if 'subtotal' in request.POST:
            subtotal_va = request.POST['subtotal']
        if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
            projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)
        if num_cus_va == 'value1':
            data = {}

            data.setdefault('1', 0)
            data.setdefault('2', 0)
            data.setdefault('3', 0)
            data.setdefault('4', 0)
            data.setdefault('5', 0)

            if projects.filter(node='AWRI'):
                data['1'] += list(projects.filter(node='AWRI').aggregate(Sum('cus_count')).values())[0]
            if projects.filter(node='Murdoch'):
                data['2'] += list(projects.filter(node='Murdoch').aggregate(Sum('cus_count')).values())[0]
            if projects.filter(node='UQ'):
                data['3'] += list(projects.filter(node='UQ').aggregate(Sum('cus_count')).values())[0]
            if projects.filter(node='UM'):
                data['4'] += list(projects.filter(node='UM').aggregate(Sum('cus_count')).values())[0]
            if projects.filter(node='UWA'):
                data['5'] += list(projects.filter(node='UWA').aggregate(Sum('cus_count')).values())[0]
            return render(request, 'allNodesByNode_num_result.html', {'AWRI': data['1'], 'Murdoch': data['2'],'UQ': data['3'],'UM': data['4'], 'UWA':data['5']})
        else:
            data = {}

            data.setdefault('1', 0)
            data.setdefault('2', 0)
            data.setdefault('3', 0)
            data.setdefault('4', 0)
            data.setdefault('5', 0)

            if projects.filter(node='AWRI'):
                data['1'] += list(projects.filter(node='AWRI').aggregate(Sum('subtotal')).values())[0]
            if projects.filter(node='Murdoch'):
                data['2'] += list(projects.filter(node='Murdoch').aggregate(Sum('subtotal')).values())[0]
            if projects.filter(node='UQ'):
                data['3'] += list(projects.filter(node='UQ').aggregate(Sum('subtotal')).values())[0]
            if projects.filter(node='UM'):
                data['4'] += list(projects.filter(node='UM').aggregate(Sum('subtotal')).values())[0]
            if projects.filter(node='UWA'):
                data['5'] += list(projects.filter(node='UWA').aggregate(Sum('subtotal')).values())[0]
            return render(request, 'allNodesByNode_total_result.html',
                          {'AWRI': data['1'], 'Murdoch': data['2'], 'UQ': data['3'], 'UM': data['4'], 'UWA': data['5']})
    # context_dict={}
    # if request.method == 'POST':
    #     from_date = request.POST['from_user_date']
    #     to_date = request.POST['to_user_date']
    #     num_cus_va = None
    #     subtotal_va = None
    #     if 'num_cus' in request.POST:
    #         num_cus_value = request.POST['num_cus']
    #     if 'subtotal' in request.POST:
    #         subtotal_value = request.POST['subtotal']
    #     if num_cus_va == 'num_cus_va':
    #         data = []
    #         Node_choice = ['AWRI', 'Murdoch', 'UQ', 'UM', 'UWA']
    #         if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
    #             projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)
    #             for t in Node_choice:
    #
    #                 if projects.filter(node=t).exists():
    #                     data.append(list(projects.filter(node=t).aggregate(Sum('subtotal')).values())[0])
    #                 else:
    #                     data.append(0)
    #         context_dict['subtotal'] = data
    #
    #         return render(request,'allNodesByNode_total_result.html',context_dict)
    #     else:
    #         data = []
    #         Node_choice = ['AWRI', 'Murdoch', 'UQ', 'UM', 'UWA']
    #         if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
    #             projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)
    #             for t in Node_choice:
    #
    #                 if projects.filter(node=t).exists():
    #                     data.append(list(projects.filter(node=t).aggregate(Sum('cus_count')).values())[0])
    #                 else:
    #                     data.append(0)
    #         context_dict['cus_count'] = data
    #         return render(request, 'allNodesByNode_num_result.html', context_dict)

def logout(request):
    if request.session.get('username', '') != '':
        del request.session['username']
    return redirect('login')

def filter_result(request,node_name_slug):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    if request.method=='POST':
        from_date = request.POST['from_user_date']
        to_date = request.POST['to_user_date']
        category_value=None
        user1_value=None
        user2_value=None
        state_value=None

        if 'Category' in request.POST:
            category_value = request.POST['Category']
        if 'User_Define1' in request.POST:
            user1_value = request.POST['User_Define1']
        if 'User_Define2' in request.POST:
            user2_value = request.POST['User_Define2']
        if 'State' in request.POST:
            state_value = request.POST['State']
        print request.POST
        print user1_value
        print user2_value
        if state_value == 'va_state':
            data = []

            if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
                projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)

                for t in STATES:
                    data.append(t)
                    if projects.filter(state=t).exists():
                        data.append(list(projects.filter(state=t).aggregate(Sum('subtotal')).values())[0])
                        data.append(list(projects.filter(state=t).aggregate(Sum('cus_count')).values())[0])
                       # data.append(projects.filter(state=t).count())
                        data.append(list(projects.filter(state=t).aggregate(Sum('num_sample')).values())[0])
                    else:
                        data.append(0)
                        data.append(0)
                        data.append(0)
                        # data.append(0)

                data = [data[x:x + 4] for x in range(0, len(data), 4)]
            else:
                data = [[0, 0, 0, 0, 0]]
            context_dict['State'] = data
        if user2_value == 'va_user_define2':
            data = []
            user2_choice = ['LIF/FOO', 'BIO', 'LIF/MED', 'ENV', 'LIF/AGR', 'LIF', 'PHA']
            if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
                projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)

                for t in user2_choice:
                    data.append(t)
                    if projects.filter(user_define2=t).exists():
                        data.append(list(projects.filter(user_define2=t).aggregate(Sum('subtotal')).values())[0])
                        data.append(list(projects.filter(user_define2=t).aggregate(Sum('cus_count')).values())[0])
                        #data.append(projects.filter(user_define2=t).count())
                        data.append(list(projects.filter(user_define2=t).aggregate(Sum('num_sample')).values())[0])
                    else:
                        data.append(0)
                        data.append(0)
                        data.append(0)
                        # data.append(0)

                data = [data[x:x + 4] for x in range(0, len(data), 4)]
            else:
                data = [[0, 0, 0, 0, 0]]
            context_dict['User_Define2'] = data
        if user1_value == 'va_user_define1':
            data = []
            user1_choice = ['PRDC', 'UNI', 'CF','CB','CO','CP','POTH']
            if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
                projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)

                for t in user1_choice:
                    data.append(t)
                    if projects.filter(user_define1=t).exists():
                        data.append(list(projects.filter(user_define1=t).aggregate(Sum('subtotal')).values())[0])
                        data.append(list(projects.filter(user_define1=t).aggregate(Sum('cus_count')).values())[0])
                        # data.append(projects.filter(user_define1=t).count())
                        data.append(list(projects.filter(user_define1=t).aggregate(Sum('num_sample')).values())[0])
                    else:
                        data.append(0)
                        data.append(0)
                        # data.append(0)
                        data.append(0)

                data = [data[x:x + 4] for x in range(0, len(data), 4)]
            else:
                data = [[0, 0, 0, 0, 0]]
            context_dict['User_Define1'] = data
        if category_value == 'va_category':
            cates = ['1', '2', '3']
            data = []
            if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
                projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)

                for t in cates:
                    data.append(t)
                    if projects.filter(category=t).exists():
                        data.append(list(projects.filter(category=t).aggregate(Sum('subtotal')).values())[0])
                        data.append(list(projects.filter(category=t).aggregate(Sum('cus_count')).values())[0])
                        # data.append(projects.filter(category=t).count())
                        data.append(list(projects.filter(category=t).aggregate(Sum('num_sample')).values())[0])
                    else:
                        data.append(0)
                        data.append(0)
                        # data.append(0)
                        data.append(0)
                data = [data[x:x + 4] for x in range(0, len(data), 4)]
            else:
                data = [[0, 0, 0, 0, 0]]
            context_dict['Category']=data
    print context_dict
    return render(request, 'nodeFilter_result.html', context_dict)




def filter_state(request):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    data = []
    if request.method == 'POST':
        from_date = request.POST['from_user_date']
        to_date = request.POST['to_user_date']


        if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
            projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)

            for t in STATES:
                data.append(t)
                if projects.filter(state=t).exists():
                    data.append(list(projects.filter(state=t).aggregate(Sum('subtotal')).values())[0])
                    data.append(list(projects.filter(state=t).aggregate(Sum('cus_count')).values())[0])
                    # data.append(projects.filter(state=t).count())
                    data.append(list(projects.filter(state=t).aggregate(Sum('num_sample')).values())[0])
                # else:
                #     data.append(0)
                #     data.append(0)
                #     data.append(0)
                #     data.append(0)

            data = [data[x:x + 4] for x in range(0, len(data), 4)]
        else:
            data = [[0, 0, 0, 0, 0]]
        context_dict['State'] = data
        return render(request, 'allNodesByState_result.html', context_dict)


def add(request):
    a=request.GET['a']
    b = request.GET['b']
    a=int(a)
    b=int(b)
    return HttpResponse(str(a+b))

def node(request, node_name_slug):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict={}
    context_dict['node_name'] = node_name_slug
    projects = Project.objects.filter(node=node_name_slug)
    context_dict['projects'] = projects
    return render(request,'node.html',context_dict)

def projects_list(request,node_name_slug):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    try:
        context_dict['node_name_slug'] = node_name_slug
        node = Node.objects.get(node_abbr=node_name_slug)
        projects = Project.objects.filter(node=node_name_slug)
        context_dict['p'] = projects
        return render(request, 'nodeProTable.html', context_dict)
    except Node.DoesNotExist:
        return render(request, 'nodeProTable.html', context_dict)






def test_projects_list(request):
    if request.session.get("username",None) ==None:
        return HttpResponseRedirect(reverse('login'))
    context_dict = {}
    try:


        p = Project.objects.filter(node='UM')
        context_dict['p'] = p


    except Node.DoesNotExist:
        pass

    return render(request,'nodeProTable.html',context_dict)


def project_detail(request,pro_id):
    if request.session.get("username",None) ==None:
        return HttpResponseRedirect(reverse('login'))
    context_dict={}
    try:
        p = Project.objects.get(pro_id=pro_id)
    except:
        p = None
    context_dict['pro_id']=p.pro_id
    context_dict['p']=p

    return render(request,'project_detail.html',context_dict)


def node_chart(request,node_name_slug):
    print request.session.get('username','no user name')
    if request.session.get('username','') =='':
        return HttpResponseRedirect('login')
    # print request.session.get('username')
    def get_aggregation(node_name_slug):
        res = {}
        res.setdefault('1',0)
        res.setdefault('2',0)
        res.setdefault('3',0)

        projects = Project.objects.filter(node=node_name_slug)

        if projects.filter(category='1'):
            res['1']+=list(projects.filter(category='1').aggregate(Sum('subtotal')).values())[0]
        if projects.filter(category='2') :
            res['2']+=list(projects.filter(category='2').aggregate(Sum('subtotal')).values())[0]
        if projects.filter(category='3') :
            res['3']+=list(projects.filter(category='3').aggregate(Sum('subtotal')).values())[0]
        return res


    data = get_aggregation(node_name_slug)

    return render(request,'nodeIncomeChart.html',{'cat1':data['1'],'cat2':data['2'],'cat3':data['3']})



# quote
# def quote_filter(request):
def quote_page(request):
    if request.session.get("username","") =="":
        return redirect("login")
    return render(request, 'quote.html')


def quote_grant(request):
    if request.session.get("username","") =="":
        return redirect("login")
    return render(request, 'quote_grant.html')

def quote_status(request):
    if request.session.get("username","") =="":
        return redirect("login")
    return render(request, 'quote_status.html')

def quote_company(request):
    if request.session.get("username","") =="":
        return redirect("login")
    return render(request, 'quote_company.html')

def quote_grant_result(request):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    if request.method == 'POST':
        from_date = request.POST['from_user_date']
        to_date = request.POST['to_user_date']
        print from_date
        arc_va = None
        n_va = None
        others_va=None
        if 'ARC' in request.POST:
            arc_va = request.POST['ARC']
        if 'N' in request.POST:
            n_va = request.POST['N']
        if 'Others' in request.POST:
            others_va = request.POST['Others']
        if quote.objects.filter(quote_date__gte=from_date, quote_date__lte=to_date).exists():
            quotes = quote.objects.filter(quote_date__gte=from_date, quote_date__lte=to_date)

            if arc_va == 'arc_va':
                ARC_qu=quotes.filter(grant_not='ARC')

                context_dict['ARC'] = ARC_qu
            if others_va == 'others_va':
                Others_q =quotes.filter(grant_not='Null')
                context_dict['others'] = Others_q
            if n_va == 'n_va':
                N_q = quotes.filter(grant_not='N')
                context_dict['N'] = N_q
            arc_sum=quotes.filter(grant_not='ARC').aggregate(Sum('quote_value')).values()[0]
            print arc_sum
            context_dict['ARC_Sum']=arc_sum
            n_sum = quotes.filter(grant_not='N').aggregate(Sum('quote_value')).values()[0]
            print n_sum
            context_dict['N_Sum'] = n_sum
            others_sum = quotes.filter(grant_not='Null').aggregate(Sum('quote_value')).values()[0]
            print others_sum
            context_dict['others_Sum'] =others_sum
            return render(request, 'quote_grant_result.html', context_dict)


def quote_grant_detail(request,quote_id):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict={}
    try:
        q = quote.objects.get(quote_id=quote_id)
    except:
        q = None
    context_dict['quote_id']=q.quote_id
    context_dict['q']=q

    return render(request,'quote_grant_detail.html',context_dict)



def quote_status_detail(request,quote_id):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict={}
    try:
        q = quote.objects.get(quote_id=quote_id)
    except:
        q = None
    context_dict['quote_id']=q.quote_id
    context_dict['q_status']=q

    return render(request,'quote_status_detail.html',context_dict)

def quote_status_result(request):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    if request.method == 'POST':
        from_date = request.POST['from_user_date']
        to_date = request.POST['to_user_date']
        print from_date
        accept_va = None
        invoice_va = None

        if 'accept' in request.POST:
            accept_va = request.POST['accept']
        if 'invoice' in request.POST:
            invoice_va= request.POST['invoice']

        if quote.objects.filter(quote_date__gte=from_date, quote_date__lte=to_date).exists():
            quotes = quote.objects.filter(quote_date__gte=from_date, quote_date__lte=to_date)

            if accept_va == 'accept_va' and  invoice_va=='invoice_va':
                q_status = quotes.filter(accept='Y').filter(invoiced_not='Y')
                Status_Sum = quotes.filter(accept='Y').filter(invoiced_not='Y').aggregate(Sum('quote_value')).values()[0]
                context_dict['q_status'] = q_status
                context_dict['Status_Sum'] = Status_Sum
            if accept_va == 'accept_va' and invoice_va != 'invoice_va':
                q_status = quotes.filter(accept='Y').filter(invoiced_not='Null')
                Status_Sum = quotes.filter(accept='Y').filter(invoiced_not='Null').aggregate(Sum('quote_value')).values()[0]
                context_dict['q_status'] = q_status
                context_dict['Status_Sum'] = Status_Sum
            if accept_va != 'accept_va' and invoice_va == 'invoice_va':
                q_status = quotes.filter(accept='Null').filter(invoiced_not = 'Y')
                Status_Sum = quotes.filter(accept='Null').filter(invoiced_not = 'Y').aggregate(Sum('quote_value')).values()[0]
                context_dict['q_status'] = q_status
                context_dict['Status_Sum'] = Status_Sum
            if accept_va != 'accept_va' and invoice_va != 'invoice_va':
                q_status = quotes.filter(accept='Null').filter(invoiced_not = 'Null')
                Status_Sum = quotes.filter(accept='Null').filter(invoiced_not='Null').aggregate(Sum('quote_value')).values()[0]
                context_dict['q_status'] = q_status
                context_dict['Status_Sum'] = Status_Sum

            return render(request, 'quote_status_result.html', context_dict)

def quote_company_result(request):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    if request.method == 'POST':

        company = request.POST['company']
        quotes = quote.objects.filter(company=company)
        context_dict['q_company']=quotes
        return render(request, 'quote_company_result.html', context_dict)

def quote_company_detail(request,quote_id):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict={}
    try:
        q = quote.objects.get(quote_id=quote_id)
    except:
        q = None
    context_dict['quote_id']=q.quote_id
    context_dict['q_company']=q

    return render(request,'quote_company_detail.html',context_dict)

def invoice_page(request):
    return render(request, 'invoice.html')

def search(request):
    return render(request, 'search.html')

def invoice_result(request):
    if request.session.get("username","") =="":
        return redirect("login")
    context_dict = {}
    if request.method=='POST':
        from_date = request.POST['from_user_date']
        to_date = request.POST['to_user_date']

        paid_va=None
        unpaid_va=None


        if 'paid' in request.POST:
            paid_va = request.POST['paid']
        if 'unpaid' in request.POST:
            unpaid_va = request.POST['unpaid']

        print paid_va
        print unpaid_va
        if paid_va == 'paid_va':


            if invoice.objects.filter(invoice_date__gte=from_date, invoice_date__lte=to_date).exists():
                invoices = invoice.objects.filter(invoice_date__gte=from_date, invoice_date__lte=to_date)
                paid=invoices.filter(paid='Y')

            context_dict['paid'] = paid
        if unpaid_va == 'unpaid_va':


            if invoice.objects.filter(invoice_date__gte=from_date, invoice_date__lte=to_date).exists():
                invoices = invoice.objects.filter(invoice_date__gte=from_date, invoice_date__lte=to_date)
                unpaid = invoices.filter(paid='N')

            context_dict['unpaid'] = unpaid
        return render(request, 'invoice_result.html', context_dict)


def invoice_detail(request,invoice_id):
    if request.session.get("username",None) ==None:
        return HttpResponseRedirect(reverse('login'))
    context_dict={}
    try:
        p = invoice.objects.get(invoice_id=invoice_id)
    except:
        p = None
    context_dict['invoice_id']=invoice_id
    context_dict['i']=p

    return render(request,'invoice_detail.html',context_dict)

def search_result(request):
    if request.session.get("username",None) ==None:
        return HttpResponseRedirect(reverse('login'))
    context_dict = {}
    if request.method == 'POST':
        from_date = request.POST['from_user_date']
        to_date = request.POST['to_user_date']
        if Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date).exists():
            projects = Project.objects.filter(pro_date__gte=from_date, pro_date__lte=to_date)
            person = request.POST['person']
            organisation = request.POST['organisation']
            service = request.POST['service']
            description = request.POST['description']
            instrument = request.POST['instrument']
            pros=projects.all()
            if person!='':
                pros=pros.filter(person=person)
            if organisation!='':
                pros=pros.filter(organization=organisation)
            if service!='':
                pros=pros.filter(service=service)
            if description!='':
                pros=pros.filter(project_name=description)
            if instrument!='':
                pros=pros.filter(instrument=instrument)
            print len(pros)
            personF=list(filter(lambda x:x.person==person if person!='' else True,pros))
            # s1=projects.filter(project_name="*")
            # s2=projects.filter(organization=organisation)
            # s3=projects.filter(service=service)
            # s4=projects.filter(instrument=instrument)
            # s5=projects.filter(person=person)

            context_dict['s'] =pros

        return render(request, 'search_result.html', context_dict)

def search_detail(request,search_id):

    if request.session.get("username", None) == None:
        return HttpResponseRedirect(reverse('login'))
    context_dict = {}
    try:
        p = Project.objects.get(pro_id=search_id)
    except:
        p = None
    context_dict['pro_id'] = search_id
    context_dict['s'] = p

    return render(request, 'search_detail.html', context_dict)


